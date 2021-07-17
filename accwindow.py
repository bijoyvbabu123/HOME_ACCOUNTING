import tkinter
from tkinter import ttk
from tkinter import messagebox
import tkcalendar
from ttkwidgets import autocomplete
import re
import os
import openpyxl
import datetime

accroot = tkinter.Tk()
accroot.minsize(width=1150, height=660)
accroot.title('Expenditure Accounting')
accroot.iconbitmap(os.path.join('files', 'acc.ico'))
# accroot.focus_set()
accroot.focus_force()

# colors if needed  and theme also............rest at the end   default color is 'SystemButtonFace'
winbg = "white smoke"
textfg = "black"

buttonbg = "gainsboro"
buttonfg = "black"

treebg = "white"
treefield = "white"
oddrowbg = "light sky blue"

messageboxbg = "white"
messageboxfg = "black"

# configuring the widgets and window
accroot.config(bg=winbg)

style = ttk.Style()
style.theme_use("clam")  # default is vista
style.configure("Treeview", background=treebg, fieldbackground=treefield, rowheight=25, font='TkDefaultFont 9 italic')
style.configure("Treeview.Heading", font='TkDefaultFont 10 bold')
style.map("Treeview", background=[('selected', 'RoyalBlue3')])

# variables for the folder name
path_for_defaults = 'Defaults'
path_for_data = 'DataEntry'

# list for storing the data
entries = []

# flag for  add or done button
iid_flag = 0

# creating folders
for i in [path_for_data, path_for_defaults]:
    if not os.path.exists(i):
        os.mkdir(i)

# creating spreadsheet for the defaults
for i in ['items.xlsx', 'category.xlsx', 'pair.xlsx', 'mode.xlsx', 'unit.xlsx']:
    if not os.path.exists(os.path.join(path_for_defaults, i)):
        workbook = openpyxl.Workbook()
        workbook.worksheets[0].title = 'data'
        workbook.save(os.path.join(path_for_defaults, i))


# focusing the next entry widget on enter key....the function to bind
def focus_next_entry(event):
    x = event.widget.tk_focusNext()
    x.focus()
    if x != buttonaddordone:
        x.select_range(0, tkinter.END)
    return "break"


# checking if any selection is made in the tree and the changing the state of edit, delete, clear
def check_tree_selection(event):
    if tree.selection():
        buttonedit.config(state=tkinter.NORMAL)
        buttondelete.config(state=tkinter.NORMAL)
        buttonclearselection.config(state=tkinter.NORMAL)
    else:
        buttonedit.config(state=tkinter.DISABLED)
        buttondelete.config(state=tkinter.DISABLED)
        buttonclearselection.config(state=tkinter.DISABLED)


# updating the tree view content after every action
def update_tree():
    totalnet = 0.00
    totalgross = 0.00

    tree.delete(*tree.get_children())

    for i in range(len(entries)):
        if i % 2 != 0:
            tree.insert(parent='', index='end', iid=i + 1, values=entries[i], tags=('oddrow',))
        else:
            tree.insert(parent='', index='end', iid=i + 1, values=entries[i])
        totalnet += float(entries[i][9])
        totalgross += float(entries[i][5]) * float(entries[i][6])

    labelnumberofentries.config(text='No. of entries = ' + str(len(entries)))

    labeltotalgross.config(text='Total Gross = ' + "{:.2f}".format(totalgross))

    labeltotalnet.config(text='Total Net = ' + "{:.2f}".format(totalnet))

    check_tree_selection(None)


# deleting an entry
def delete_entry():
    x = int(tree.selection()[0])
    del entries[x - 1]
    update_tree()


# editing an entry
def edit_entry():
    global iid_flag
    iid_flag = int(tree.selection()[0])

    buttonaddordone.config(text='DONE')

    datepicker.delete(0, tkinter.END)
    itementry.delete(0, tkinter.END)
    categoryentry.delete(0, tkinter.END)
    modeentry.delete(0, tkinter.END)
    unitentry.delete(0, tkinter.END)
    quantityentry.delete(0, tkinter.END)
    rateentry.delete(0, tkinter.END)
    disallentry.delete(0, tkinter.END)
    disrecentry.delete(0, tkinter.END)
    netentry.delete(0, tkinter.END)

    datepicker.insert(0, entries[iid_flag - 1][0])
    itementry.insert(0, entries[iid_flag - 1][1])
    categoryentry.insert(0, entries[iid_flag - 1][2])
    modeentry.insert(0, entries[iid_flag - 1][3])
    unitentry.insert(0, entries[iid_flag - 1][4])
    quantityentry.insert(0, entries[iid_flag - 1][5])
    rateentry.insert(0, entries[iid_flag - 1][6])
    disallentry.insert(0, entries[iid_flag - 1][7])
    disrecentry.insert(0, entries[iid_flag - 1][8])
    netentry.insert(0, entries[iid_flag - 1][9])

    datepicker.focus_set()


# exit without saving
def exit_without_saving():
    global entries
    del entries
    accroot.destroy()


# save and exit
def save_exit():
    buttonaddordone.config(state=tkinter.DISABLED)
    buttonnosaveexit.config(state=tkinter.DISABLED)
    buttonsaveexit.config(state=tkinter.DISABLED)

    pb = ttk.Progressbar(accroot, mode='determinate')
    pb.place(relwidth=0.75, relx=0.22, rely=0.965)

    total_entries = len(entries)

    while entries:
        file = entries[0][0][-4:] + '.xlsx'

        # data entering part
        datawb = openpyxl.load_workbook(os.path.join(path_for_data, file))
        dataws = datawb['data']

        if dataws.max_row == 1:
            for i in range(1, 11):
                dataws.cell(row=2, column=i).value = entries[0][i - 1]
        else:
            pos = 0
            for i in range(dataws.max_row, 1, -1):
                if datetime.datetime.strptime(entries[0][0], '%d/%m/%Y') >= datetime.datetime.strptime(
                        dataws.cell(column=1, row=i).value, '%d/%m/%Y'):
                    pos = i + 1
                    break
            if pos == 0:
                pos = 2
            dataws.insert_rows(pos)
            for i in range(1, 11):
                dataws.cell(row=pos, column=i).value = entries[0][i - 1]

        datawb.save(os.path.join(path_for_data, file))

        # items defaults entering part
        itemswb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'items.xlsx'))
        itemsws = itemswb['data']

        itemfound = 0
        for i in itemsws['A']:
            if i.value == entries[0][1]:
                itemfound = 1
                break

        if not itemfound:

            if itemsws.max_row == 1 and itemsws.cell(row=1, column=1).value is None:
                itemsws.cell(column=1, row=1).value = entries[0][1]
            else:
                pos = 1
                for i in range(itemsws.max_row, 0, -1):
                    if entries[0][1] > itemsws.cell(column=1, row=i).value:
                        pos = i + 1
                        break
                itemsws.insert_rows(pos)
                itemsws.cell(column=1, row=pos).value = entries[0][1]

        itemswb.save(os.path.join(path_for_defaults, 'items.xlsx'))

        # category defaults entering part
        categorywb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'category.xlsx'))
        categoryws = categorywb['data']

        categoryfound = 0
        for i in categoryws['A']:
            if i.value == entries[0][2]:
                categoryfound = 1
                break

        if not categoryfound:

            if categoryws.max_row == 1 and categoryws.cell(row=1, column=1).value is None:
                categoryws.cell(column=1, row=1).value = entries[0][2]
            else:
                pos = 1
                for i in range(categoryws.max_row, 0, -1):
                    if entries[0][2] > categoryws.cell(column=1, row=i).value:
                        pos = i + 1
                        break
                categoryws.insert_rows(pos)
                categoryws.cell(column=1, row=pos).value = entries[0][2]

        categorywb.save(os.path.join(path_for_defaults, 'category.xlsx'))

        # mode defaults entering part
        modewb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'mode.xlsx'))
        modews = modewb['data']

        modefound = 0
        for i in modews['A']:
            if i.value == entries[0][3]:
                modefound = 1
                break

        if not modefound:

            if modews.max_row == 1 and modews.cell(row=1, column=1).value is None:
                modews.cell(column=1, row=1).value = entries[0][3]
            else:
                pos = 1
                for i in range(modews.max_row, 0, -1):
                    if entries[0][3] > modews.cell(column=1, row=i).value:
                        pos = i + 1
                        break
                modews.insert_rows(pos)
                modews.cell(column=1, row=pos).value = entries[0][3]

        modewb.save(os.path.join(path_for_defaults, 'mode.xlsx'))

        # unit defaults entering part
        unitwb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'unit.xlsx'))
        unitws = unitwb['data']

        unitfound = 0
        for i in unitws['A']:
            if i.value == entries[0][4]:
                unitfound = 1
                break

        if not unitfound:

            if unitws.max_row == 1 and unitws.cell(row=1, column=1).value is None:
                unitws.cell(column=1, row=1).value = entries[0][4]
            else:
                pos = 1
                for i in range(unitws.max_row, 0, -1):
                    if entries[0][4] > unitws.cell(column=1, row=i).value:
                        pos = i + 1
                        break
                unitws.insert_rows(pos)
                unitws.cell(column=1, row=pos).value = entries[0][4]

        unitwb.save(os.path.join(path_for_defaults, 'unit.xlsx'))

        # pair entering part
        pairwb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'pair.xlsx'))
        pairws = pairwb['data']

        pairfound = 0
        for i in pairws['A']:
            if i.value == entries[0][1]:
                pairfound = 1
                break

        if not pairfound:

            if pairws.max_row == 1 and pairws.cell(row=1, column=1).value is None:
                pairws.cell(column=1, row=1).value = entries[0][1]
                pairws.cell(column=2, row=1).value = entries[0][2]
            else:
                pos = 1
                for i in range(pairws.max_row, 0, -1):
                    if entries[0][1] > pairws.cell(column=1, row=i).value:
                        pos = i + 1
                        break
                pairws.insert_rows(pos)
                pairws.cell(column=1, row=pos).value = entries[0][1]
                pairws.cell(column=2, row=pos).value = entries[0][2]

        pairwb.save(os.path.join(path_for_defaults, 'pair.xlsx'))

        del entries[0]
        pb['value'] = ((total_entries - len(entries)) / total_entries) * 100
        accroot.update()

    accroot.destroy()


# always title type entry in items, category, mode, unit
def always_title(event, a):
    x = a.get()
    a.delete(0, tkinter.END)
    a.insert(0, x.title().strip())
    del x


# item category auto fill
def category_auto_fill(event):
    always_title(event=None, a=itementry)

    ws = openpyxl.load_workbook(os.path.join(path_for_defaults, 'pair.xlsx'))['data']
    for i in range(1, ws.max_row + 1):
        if ws.cell(column=1, row=i).value == itementry.get():
            categoryentry.delete(0, tkinter.END)
            categoryentry.insert(0, ws.cell(column=2, row=i).value)
            modeentry.focus_set()
            break


# validating the input in qty, rate, dis all, dis rec, net as only numerical or null
def only_numeric(letter):
    if bool(re.match('\d*\.\d*$', letter)):
        return True
    if letter.isdigit():
        return True
    if letter == '':
        return True
    else:
        return False


# converting all the numeric values in the entry to float for qty, rate, dis all
def entry_to_float(event, a):
    if a.get() != '':
        x = float(a.get())
        a.delete(0, tkinter.END)

        if a == quantityentry:
            a.insert(0, "{:.3f}".format(x))
        else:
            a.insert(0, "{:.2f}".format(x))

        if a == quantityentry or a == rateentry:
            if x == 0:
                a.delete(0, tkinter.END)
        del x


# after the dis rec looses focus its value is converted to float and net is calculated if possible
def after_disrec(event):
    if disrecentry.get() != '':
        x = float(disrecentry.get())
        disrecentry.delete(0, tkinter.END)
        disrecentry.insert(0, "{:.2f}".format(x))
        del x

    if (quantityentry.get() != '') and (rateentry.get() != '') and (disallentry.get() != '') and (
            disrecentry.get() != ''):
        net = (float(quantityentry.get()) * float(rateentry.get())) + float(disallentry.get()) - float(
            disrecentry.get())
        netentry.delete(0, tkinter.END)
        netentry.insert(0, "{:.2f}".format(net))
        del net


# checking if any of the fields among items, category, mode and unit are empty
def no_empty_entry():
    if itementry.get() == '':
        messagebox.showwarning('Should not be Empty', "Please don't leave the 'Items' field Empty")
        itementry.focus_set()
    elif datepicker.get() == '':
        messagebox.showwarning('Should not be Empty', "Please don't leave the 'Date' field Empty")
        datepicker.focus_set()
    elif categoryentry.get() == '':
        messagebox.showwarning('Should not be Empty', "Please don't leave the 'Category' field Empty")
        categoryentry.focus_set()
    elif modeentry.get() == '':
        messagebox.showwarning('Should not be Empty', "Please don't leave the 'Mode' field Empty")
        modeentry.focus_set()
    elif unitentry.get() == '':
        messagebox.showwarning('Should not be Empty', "Please don't leave the 'Unit' field Empty")
        unitentry.focus_set()


# converting the net entry to float and calculating if any missing fields and calling no_empty_entry
def after_net(event):
    if netentry.get() != '':
        x = float(netentry.get())
        netentry.delete(0, tkinter.END)
        netentry.insert(0, "{:.2f}".format(x))
        del x

    if (quantityentry.get() == '') and (rateentry.get() != '') and (disallentry.get() != '') and \
            (disrecentry.get() != '') and (netentry.get() != ''):
        qty = (float(netentry.get()) - float(disallentry.get()) + float(disrecentry.get())) / float(rateentry.get())
        quantityentry.delete(0, tkinter.END)
        quantityentry.insert(0, "{:.3f}".format(qty))
        del qty

    elif (quantityentry.get() != '') and (rateentry.get() == '') and (disallentry.get() != '') and \
            (disrecentry.get() != '') and (netentry.get() != ''):
        r = (float(netentry.get()) - float(disallentry.get()) + float(disrecentry.get())) / float(quantityentry.get())
        rateentry.delete(0, tkinter.END)
        rateentry.insert(0, "{:.2f}".format(r))
        del r

    elif (quantityentry.get() != '') and (rateentry.get() != '') and (disallentry.get() == '') and \
            (disrecentry.get() != '') and (netentry.get() != ''):
        da = float(netentry.get()) + float(disrecentry.get()) - (
                float(quantityentry.get()) * float(rateentry.get()))
        disallentry.delete(0, tkinter.END)
        disallentry.insert(0, "{:.2f}".format(da))
        del da

    elif (quantityentry.get() != '') and (rateentry.get() != '') and (disallentry.get() != '') and \
            (disrecentry.get() == '') and (netentry.get() != ''):
        dr = (float(quantityentry.get()) * float(rateentry.get())) + float(disallentry.get()) - float(
            netentry.get())
        disrecentry.delete(0, tkinter.END)
        disrecentry.insert(0, "{:.2f}".format(dr))
        del dr

    else:
        if quantityentry.get() == '':
            messagebox.showwarning('Not Enough Data', 'Not more than ONE among Quantity, Rate, Discount Allowed, \
Discount Received and Net Price can be Empty')
            quantityentry.focus_set()
        elif rateentry.get() == '':
            messagebox.showwarning('Not Enough Data', 'Not more than ONE among Quantity, Rate, Discount Allowed, \
Discount Received and Net Price can be Empty')
            rateentry.focus_set()
        elif disallentry.get() == '':
            messagebox.showwarning('Not Enough Data', 'Not more than ONE among Quantity, Rate, Discount Allowed, \
Discount Received and Net Price can be Empty')
            disallentry.focus_set()
        elif disrecentry.get() == '':
            messagebox.showwarning('Not Enough Data', 'Not more than ONE among Quantity, Rate, Discount Allowed, \
Discount Received and Net Price can be Empty')
            disrecentry.focus_set()
        elif netentry.get() == '':
            messagebox.showwarning('Not Enough Data', 'Not more than ONE among Quantity, Rate, Discount Allowed, \
Discount Received and Net Price can be Empty')
            netentry.focus_set()


# adding to the tree view or editing a tree view content on button add or done
def add_or_edit_entry():
    if itementry.get() != '':
        category_auto_fill(None)
    if itementry.get() == '' or categoryentry.get() == '' or modeentry.get() == '' or unitentry.get() == '' or datepicker.get() == '':
        no_empty_entry()
    always_title(None, categoryentry)
    always_title(None, modeentry)
    always_title(None, unitentry)
    entry_to_float(None, quantityentry)
    entry_to_float(None, rateentry)
    entry_to_float(None, disallentry)
    after_disrec(None)
    after_net(None)

    if datepicker.get() != '' and itementry.get() != '' and categoryentry.get() != '' and modeentry.get() != '' and \
            unitentry.get() != '' and quantityentry.get() != '' and rateentry.get() != '' and disallentry.get() != '' \
            and disrecentry.get() != '' and netentry.get() != '':

        global iid_flag

        current_entry = []

        file = datepicker.get()[-4:] + '.xlsx'
        if not os.path.exists(os.path.join(path_for_data, file)):
            workbook = openpyxl.Workbook()
            workbook.worksheets[0].title = 'data'
            workbook.save(os.path.join(path_for_data, file))
            wb = openpyxl.load_workbook(os.path.join(path_for_data, file))
            ws = wb['data']
            ws.cell(row=1, column=1).value = 'DATE'
            ws.cell(row=1, column=2).value = 'ITEM'
            ws.cell(row=1, column=3).value = 'CATEGORY'
            ws.cell(row=1, column=4).value = 'MODE'
            ws.cell(row=1, column=5).value = 'UNIT'
            ws.cell(row=1, column=6).value = 'QUANTITY'
            ws.cell(row=1, column=7).value = 'RATE'
            ws.cell(row=1, column=8).value = 'DIS. ALL'
            ws.cell(row=1, column=9).value = 'DIS. REC'
            ws.cell(row=1, column=10).value = 'NET'
            wb.save(os.path.join(path_for_data, file))

        current_entry.append(datepicker.get())
        current_entry.append(itementry.get())
        current_entry.append(categoryentry.get())
        current_entry.append(modeentry.get())
        current_entry.append(unitentry.get())
        current_entry.append(quantityentry.get())
        current_entry.append(rateentry.get())
        current_entry.append(disallentry.get())
        current_entry.append(disrecentry.get())
        current_entry.append(netentry.get())

        if iid_flag != 0:
            entries[int(iid_flag) - 1] = current_entry
        else:
            entries.append(current_entry)

        iid_flag = 0

        del current_entry

        itementry.delete(0, tkinter.END)
        categoryentry.delete(0, tkinter.END)
        modeentry.delete(0, tkinter.END)
        unitentry.delete(0, tkinter.END)
        quantityentry.delete(0, tkinter.END)
        quantityentry.insert(0, 0.0)
        rateentry.delete(0, tkinter.END)
        rateentry.insert(0, 0.0)
        disallentry.delete(0, tkinter.END)
        disallentry.insert(0, 0.0)
        disrecentry.delete(0, tkinter.END)
        disrecentry.insert(0, 0.0)
        netentry.delete(0, tkinter.END)
        netentry.insert(0, 0.0)

        buttonaddordone.config(text='ADD')

        update_tree()

        itementry.focus_set()


# assigning completion list for item, category, mode, entry on getting focus
def set_completion(event, a):
    l = []
    ws = None

    if a == modeentry:
        ws = openpyxl.load_workbook(os.path.join(path_for_defaults, 'mode.xlsx'))['data']
    elif a == unitentry:
        ws = openpyxl.load_workbook(os.path.join(path_for_defaults, 'unit.xlsx'))['data']
    elif a == itementry:
        ws = openpyxl.load_workbook(os.path.join(path_for_defaults, 'items.xlsx'))['data']
    elif a == categoryentry:
        ws = openpyxl.load_workbook(os.path.join(path_for_defaults, 'category.xlsx'))['data']

    for i in ws['A']:
        l.append(i.value)
    if l == [None]:
        l = []

    a.set_completion_list(l)


# on pressing the button item list or button category list
def list_window(a):
    if a == itementry:
        itementry.focus_set()
    if a == categoryentry:
        categoryentry.focus_set()
    window = tkinter.Toplevel()
    window.config(bg=winbg)
    window.focus_set()
    ws = None
    wb = None
    if a == itementry:
        window.title('Item List')
        wb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'items.xlsx'))
        ws = wb['data']
    elif a == categoryentry:
        window.title('Category List')
        wb = openpyxl.load_workbook(os.path.join(path_for_defaults, 'category.xlsx'))
        ws = wb['data']
    window.focus_set()
    window.resizable(width=False, height=False)
    window.minsize(width=250, height=300)

    def update_list(event):
        l.delete(0, tkinter.END)
        if e.get() == '':
            b.config(text='select')
            for i in ws['A']:
                l.insert(tkinter.END, i.value)
        else:
            if ws['A'][0].value is not None:
                for i in ws['A']:
                    if e.get().lower() in i.value.lower():
                        l.insert(tkinter.END, i.value)

        if a == categoryentry and e.get() != '':
            for i in ws['A']:
                if ws['A'][0].value is not None:
                    if e.get().lower().strip() == i.value.lower():
                        b.config(text='select')
                        break
                b.config(text='ADD')

    def on_select(event):
        if l.curselection():
            if a == itementry:
                itementry.delete(0, tkinter.END)
                itementry.insert(0, l.get(l.curselection()))
            elif a == categoryentry:
                categoryentry.delete(0, tkinter.END)
                categoryentry.insert(0, l.get(l.curselection()))

        elif b.cget('text') == 'ADD':
            yesorno = messagebox.askyesno("Add the 'category' ?", "Do you want to add the entered Category '"
                                          + e.get().title().strip() + "' to the available category list ?")
            if yesorno:
                x = e.get().title().strip()
                categoryentry.delete(0, tkinter.END)
                categoryentry.insert(0, x)

                categoryfound = 0
                for i in ws['A']:
                    if i.value == x:
                        categoryfound = 1
                        break

                if not categoryfound:

                    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
                        ws.cell(column=1, row=1).value = x
                    else:
                        pos = 1
                        for i in range(ws.max_row, 0, -1):
                            if x > ws.cell(column=1, row=i).value:
                                pos = i + 1
                                break
                        ws.insert_rows(pos)
                        ws.cell(column=1, row=pos).value = x

        if a == itementry:
            wb.save(os.path.join(path_for_defaults, 'items.xlsx'))
        if a == categoryentry:
            wb.save(os.path.join(path_for_defaults, 'category.xlsx'))

        set_completion(None, categoryentry)
        a.focus_set()
        window.destroy()

    def on_click(event):
        e.delete(0, tkinter.END)
        if l.curselection():
            e.insert(0, l.get(l.curselection()))
        b.config(text='select')

    e = tkinter.Entry(window)
    e.place(relwidth=7 / 10, relx=0.05, rely=0.03, relheight=0.08)
    e.focus_set()
    e.bind('<FocusIn>', update_list)
    e.bind('<KeyRelease>', update_list)

    s = tkinter.Scrollbar(window, orient=tkinter.VERTICAL)
    s.place(relx=0.83, relheight=0.83, rely=0.14)

    l = tkinter.Listbox(window, yscrollcommand=s.set)
    l.place(relx=0.03, rely=0.14, relheight=0.83, relwidth=0.8)
    l.bind('<Double-Button-1>', on_select)
    l.bind('<Return>', on_select)
    l.bind('<ButtonRelease-1>', on_click)
    l.bind('<KeyRelease>', on_click)

    s.config(command=l.yview)

    b = tkinter.Button(window, text='select', command=lambda event=None: on_select(event))
    b.place(rely=0.03, relwidth=2 / 10, relx=0.77, relheight=0.08)

    if a == itementry:
        wb.save(os.path.join(path_for_defaults, 'items.xlsx'))
    if a == categoryentry:
        wb.save(os.path.join(path_for_defaults, 'category.xlsx'))

    set_completion(None, categoryentry)

    window.mainloop()


# showing information
def show_info(event, a):
    if a == datepicker:
        info.place(relx=0.175, rely=0.0256)
        info.config(text="Enter the Date of Purchase in the format DD/MM/YYYY")
    elif a == itementry:
        info.place(relx=0.205, rely=0.0256 * 5)
        info.config(text="Name of the Item or Service you paid for")
    elif a == categoryentry:
        info.place(relx=0.205, rely=0.0256 * 9)
        info.config(text="Category in which the current Item/Service belongs (eg:Apple belongs to the category Fruits)")
    elif a == modeentry:
        info.place(relx=0.175, rely=0.0256 * 13)
        info.config(text="Mode by which the Payment was done (Cash, Credit Card, Net Banking, etc)")
    elif a == unitentry:
        info.place(relx=0.175, rely=0.0256 * 17)
        info.config(text="Unit in which the current Item is expressed (If it is 3 Kilogram of Apple, Kilogram is the "
                         "Unit) (If it is 3 Pencils, 'Numbers' can be the Unit)")
    elif a == quantityentry:
        info.place(relx=0.175, rely=0.0256 * 21)
        info.config(text="How much of the Item did you Purchase ?")
    elif a == rateentry:
        info.place(relx=0.175, rely=0.0256 * 25)
        info.config(text="Cost for 1 Unit of the Item")
    elif a == disallentry:
        info.place(relx=0.175, rely=0.0256 * 29)
        info.config(text="It is the Discount you allowed to the Vendor. If the actual cost was 75 and you paid 80, "
                         "then you have allowed a discount of 5. (It is actually a Lose for you)")
    elif a == disrecentry:
        info.place(relx=0.175, rely=0.0256 * 29)
        info.config(text="It is the Discount you received. If the actual cost was 75 and you had to pay only 70, "
                         "then you received a discount of 5")
    elif a == netentry:
        info.place(relx=0.175, rely=0.0256 * 33)
        info.config(text="It is the Final Amount you Paid for the current Item/Service")


# hiding information
def hide_info(event):
    info.place_forget()


datepicker = tkcalendar.DateEntry(accroot, date_pattern='dd/mm/yyyy')
datepicker.place(relx=0.02, rely=0.0256, relwidth=0.15)
datepicker.focus_set()
datepicker.bind("<Return>", focus_next_entry)
datepicker.bind("<Enter>", lambda event, a=datepicker: show_info(event, a))
datepicker.bind("<Leave>", hide_info)

labelitem = tkinter.Label(accroot, text='Item')
labelitem.place(relx=0.02, rely=0.0256 * 4 - 0.0065)
itementry = autocomplete.AutocompleteEntry(accroot, completevalues=[])
itementry.place(relx=0.02, rely=0.0256 * 5, relwidth=0.15)
labelitem.bind("<Enter>", lambda event, a=itementry: show_info(event, a))
itementry.bind("<Enter>", lambda event, a=itementry: show_info(event, a))
labelitem.bind("<Leave>", hide_info)
itementry.bind("<Leave>", hide_info)
itementry.bind("<FocusOut>", category_auto_fill)
itementry.bind("<Return>", focus_next_entry)

labelcategory = tkinter.Label(accroot, text='Category')
labelcategory.place(relx=0.02, rely=0.0256 * 8 - 0.0065)
categoryentry = autocomplete.AutocompleteEntry(accroot, completevalues=[])
categoryentry.place(relx=0.02, rely=0.0256 * 9, relwidth=0.15)
labelcategory.bind("<Enter>", lambda event, a=categoryentry: show_info(event, a))
categoryentry.bind("<Enter>", lambda event, a=categoryentry: show_info(event, a))
labelcategory.bind("<Leave>", hide_info)
categoryentry.bind("<Leave>", hide_info)
categoryentry.bind("<FocusOut>", lambda event, a=categoryentry: always_title(event, a))
categoryentry.bind("<Return>", focus_next_entry)

labelmode = tkinter.Label(accroot, text='Mode Of Payment')
labelmode.place(relx=0.02, rely=0.0256 * 12 - 0.0065)
modeentry = autocomplete.AutocompleteCombobox(accroot, completevalues=[])
modeentry.place(relx=0.02, rely=0.0256 * 13, relwidth=0.15)
labelmode.bind("<Enter>", lambda event, a=modeentry: show_info(event, a))
modeentry.bind("<Enter>", lambda event, a=modeentry: show_info(event, a))
labelmode.bind("<Leave>", hide_info)
modeentry.bind("<Leave>", hide_info)
modeentry.bind("<FocusOut>", lambda event, a=modeentry: always_title(event, a))
modeentry.bind("<Return>", focus_next_entry)

labelunit = tkinter.Label(accroot, text='Unit (Quantity Description)')
labelunit.place(relx=0.02, rely=0.0256 * 16 - 0.0065)
unitentry = autocomplete.AutocompleteCombobox(accroot, completevalues=[])
unitentry.place(relx=0.02, rely=0.0256 * 17, relwidth=0.15)
labelunit.bind("<Enter>", lambda event, a=unitentry: show_info(event, a))
unitentry.bind("<Enter>", lambda event, a=unitentry: show_info(event, a))
labelunit.bind("<Leave>", hide_info)
unitentry.bind("<Leave>", hide_info)
unitentry.bind("<FocusOut>", lambda event, a=unitentry: always_title(event, a))
unitentry.bind("<Return>", focus_next_entry)

labelquantity = tkinter.Label(accroot, text='Quantity')
labelquantity.place(relx=0.02, rely=0.0256 * 20 - 0.0065)
quantityentry = tkinter.Entry(accroot)
quantityentry.place(relx=0.02, rely=0.0256 * 21, relwidth=0.15)
labelquantity.bind("<Enter>", lambda event, a=quantityentry: show_info(event, a))
quantityentry.bind("<Enter>", lambda event, a=quantityentry: show_info(event, a))
labelquantity.bind("<Leave>", hide_info)
quantityentry.bind("<Leave>", hide_info)
quantityentry.insert(0, 0.0)
quantityentry.config(validate='all', validatecommand=(accroot.register(only_numeric), '%P'))
quantityentry.bind("<FocusOut>", lambda event, a=quantityentry: entry_to_float(event, a))
quantityentry.bind("<Return>", focus_next_entry)

labelrate = tkinter.Label(accroot, text='Rate')
labelrate.place(relx=0.02, rely=0.0256 * 24 - 0.0065)
rateentry = tkinter.Entry(accroot)
rateentry.place(relx=0.02, rely=0.0256 * 25, relwidth=0.15)
labelrate.bind("<Enter>", lambda event, a=rateentry: show_info(event, a))
rateentry.bind("<Enter>", lambda event, a=rateentry: show_info(event, a))
labelrate.bind("<Leave>", hide_info)
rateentry.bind("<Leave>", hide_info)
rateentry.insert(0, 0.0)
rateentry.config(validate='all', validatecommand=(accroot.register(only_numeric), '%P'))
rateentry.bind("<FocusOut>", lambda event, a=rateentry: entry_to_float(event, a))
rateentry.bind("<Return>", focus_next_entry)

labeldisall = tkinter.Label(accroot, text='Dis. Allowed')
labeldisall.place(relx=0.02, rely=0.0256 * 28 - 0.0065)
disallentry = tkinter.Entry(accroot)
disallentry.place(relx=0.02, rely=0.0256 * 29, relwidth=0.07)
labeldisall.bind("<Enter>", lambda event, a=disallentry: show_info(event, a))
disallentry.bind("<Enter>", lambda event, a=disallentry: show_info(event, a))
labeldisall.bind("<Leave>", hide_info)
disallentry.bind("<Leave>", hide_info)
disallentry.insert(0, 0.0)
disallentry.config(validate='all', validatecommand=(accroot.register(only_numeric), '%P'))
disallentry.bind("<FocusOut>", lambda event, a=disallentry: entry_to_float(event, a))
disallentry.bind("<Return>", focus_next_entry)

labeldisrec = tkinter.Label(accroot, text='Dis. Received')
labeldisrec.place(relx=0.02 + 0.07 + 0.01, rely=0.0256 * 28 - 0.0065)
disrecentry = tkinter.Entry(accroot)
disrecentry.place(relx=0.02 + 0.07 + 0.01, rely=0.0256 * 29, relwidth=0.07)
labeldisrec.bind("<Enter>", lambda event, a=disrecentry: show_info(event, a))
disrecentry.bind("<Enter>", lambda event, a=disrecentry: show_info(event, a))
labeldisrec.bind("<Leave>", hide_info)
disrecentry.bind("<Leave>", hide_info)
disrecentry.insert(0, 0.0)
disrecentry.config(validate='all', validatecommand=(accroot.register(only_numeric), '%P'))
disrecentry.bind("<FocusOut>", after_disrec)
disrecentry.bind("<Return>", focus_next_entry)

labelnet = tkinter.Label(accroot, text='Net Price')
labelnet.place(relx=0.02, rely=0.0256 * 32 - 0.0065)
netentry = tkinter.Entry(accroot)
netentry.place(relx=0.02, rely=0.0256 * 33, relwidth=0.15)
labelnet.bind("<Enter>", lambda event, a=netentry: show_info(event, a))
netentry.bind("<Enter>", lambda event, a=netentry: show_info(event, a))
labelnet.bind("<Leave>", hide_info)
netentry.bind("<Leave>", hide_info)
netentry.insert(0, 0.0)
netentry.config(validate='all', validatecommand=(accroot.register(only_numeric), '%P'))
netentry.bind("<FocusOut>", after_net)
netentry.bind("<Return>", focus_next_entry)

buttonaddordone = tkinter.Button(accroot, text='ADD', command=add_or_edit_entry, relief='groove')
buttonaddordone.place(relx=0.045, rely=0.0256 * 36, relheight=0.0256 * 2, relwidth=0.10)
buttonaddordone.bind("<Return>", lambda event=None: buttonaddordone.invoke())
buttonaddordone.bind("<Enter>", lambda event=None: buttonaddordone.config(bg="SteelBlue1"))
buttonaddordone.bind("<FocusIn>", lambda event=None: buttonaddordone.config(bg="SteelBlue1"))
buttonaddordone.bind("<Leave>", lambda event=None: buttonaddordone.config(bg=buttonbg))
buttonaddordone.bind("<FocusOut>", lambda event=None: buttonaddordone.config(bg=buttonbg))

buttonsaveexit = tkinter.Button(accroot, text='SAVE and EXIT', command=save_exit, relief='groove')
buttonsaveexit.place(relwidth=0.15, relheight=0.0256 * 3, relx=0.425, rely=0.870)
buttonsaveexit.bind("<Return>", lambda event=None: buttonsaveexit.invoke())
buttonsaveexit.bind("<Enter>", lambda event=None: buttonsaveexit.config(bg="green2"))
buttonsaveexit.bind("<FocusIn>", lambda event=None: buttonsaveexit.config(bg="green2"))
buttonsaveexit.bind("<Leave>", lambda event=None: buttonsaveexit.config(bg=buttonbg))
buttonsaveexit.bind("<FocusOut>", lambda event=None: buttonsaveexit.config(bg=buttonbg))

buttonnosaveexit = tkinter.Button(accroot, text='EXIT without SAVING', command=exit_without_saving, relief='groove')
buttonnosaveexit.place(relwidth=0.15, relheight=0.0256 * 3, relx=0.615, rely=0.870)
buttonnosaveexit.bind("<Return>", lambda event=None: buttonnosaveexit.invoke())
buttonnosaveexit.bind("<Enter>", lambda event=None: buttonnosaveexit.config(bg="red"))
buttonnosaveexit.bind("<FocusIn>", lambda event=None: buttonnosaveexit.config(bg="red"))
buttonnosaveexit.bind("<Leave>", lambda event=None: buttonnosaveexit.config(bg=buttonbg))
buttonnosaveexit.bind("<FocusOut>", lambda event=None: buttonnosaveexit.config(bg=buttonbg))

buttonedit = tkinter.Button(accroot, text='Edit', state=tkinter.DISABLED, command=edit_entry)
buttonedit.place(relwidth=0.07, relheight=0.0256 * 2, rely=0.79, relx=0.47)

buttondelete = tkinter.Button(accroot, text='Delete', state=tkinter.DISABLED, command=delete_entry)
buttondelete.place(relwidth=0.07, relheight=0.0256 * 2, rely=0.79, relx=0.56)

buttonclearselection = tkinter.Button(accroot, text='Clear Selection', state=tkinter.DISABLED, command=update_tree)
buttonclearselection.place(relwidth=0.07, relheight=0.0256 * 2, rely=0.79, relx=0.65)

buttonitemlist = tkinter.Button(accroot, text='..', command=lambda a=itementry: list_window(a))
buttonitemlist.place(relx=0.18, rely=0.0256 * 5, height=21, relwidth=0.02)
buttonitemlist.bind("<Return>", lambda event=None: buttonitemlist.invoke())

buttoncategorylist = tkinter.Button(accroot, text='..', command=lambda a=categoryentry: list_window(a))
buttoncategorylist.place(relx=0.18, rely=0.0256 * 9, height=21, relwidth=0.02)
buttoncategorylist.bind("<Return>", lambda event=None: buttoncategorylist.invoke())

labelnumberofentries = tkinter.Label(accroot, text='No. of entries = 0')
labelnumberofentries.place(relx=0.22, rely=0.73)

labeltotalgross = tkinter.Label(accroot, text='Total Gross = 0.00')
labeltotalgross.place(relx=0.63, rely=0.73)

labeltotalnet = tkinter.Label(accroot, text='Total Net = 0.00')
labeltotalnet.place(relx=0.865, rely=0.73)

treescroll = tkinter.Scrollbar(accroot, orient=tkinter.VERTICAL)
treescroll.place(relheight=0.7, relx=0.97, rely=0.03)

tree = ttk.Treeview(accroot, show=['headings'], yscrollcommand=treescroll.set, selectmode='browse')
tree.place(relwidth=0.75, relheight=0.7, relx=0.22, rely=0.03)
tree.bind('<FocusIn>', check_tree_selection)
tree.bind('<ButtonRelease-1>', check_tree_selection)
tree.bind('<KeyRelease>', check_tree_selection)

treescroll.config(command=tree.yview)

tree['columns'] = ('date', 'item', 'cat', 'mode', 'unit', 'qty', 'rate', 'da', 'dr', 'net')

accroot.update()
tree.update()
tree.column('date', anchor=tkinter.CENTER, width=int(tree.winfo_width() / 10))
tree.column('item', anchor=tkinter.W, width=int(tree.winfo_width() / 8))
tree.column('cat', anchor=tkinter.W, width=int(tree.winfo_width() / 7.5))
tree.column('mode', anchor=tkinter.W, width=int(tree.winfo_width() / 9))
tree.column('unit', anchor=tkinter.W, width=int(tree.winfo_width() / 12))
tree.column('qty', anchor=tkinter.E, width=int(tree.winfo_width() / 11))
tree.column('rate', anchor=tkinter.E, width=int(tree.winfo_width() / 11))
tree.column('da', anchor=tkinter.E, width=int(tree.winfo_width() / 13))
tree.column('dr', anchor=tkinter.E, width=int(tree.winfo_width() / 13))
tree.column('net', anchor=tkinter.E, width=int(tree.winfo_width() / 8.5))
tree.heading('date', text='Date')  # , anchor=tkinter.W
tree.heading('item', text='Item')
tree.heading('cat', text='Category')
tree.heading('mode', text='Mode')
tree.heading('unit', text='Unit')
tree.heading('qty', text='Quantity')
tree.heading('rate', text='Rate')
tree.heading('da', text='Dis. All')
tree.heading('dr', text='Dis. Rec')
tree.heading('net', text='Net Price')

update_tree()

tree.tag_configure('oddrow', background=oddrowbg)

set_completion(None, itementry)
set_completion(None, categoryentry)
set_completion(None, modeentry)
set_completion(None, unitentry)

info = tkinter.Message(accroot, text="", bg=messageboxbg, foreground=messageboxfg, font='TkDefaultFont 10 italic')

# editing all the labels
for i in [labelitem, labelcategory, labelmode, labelunit, labelquantity, labelrate, labeldisall, labeldisrec, labelnet,
          labelnumberofentries, labeltotalnet, labeltotalgross]:
    i.config(bg=winbg, fg=textfg, font='TkDefaultFont 10 italic')

# editing main buttons
for i in [buttonaddordone, buttonsaveexit, buttonnosaveexit, buttonitemlist, buttoncategorylist]:
    i.config(bg=buttonbg, fg=buttonfg, font='TkDefaultFont 11 bold italic')

# editing three tree buttons
for i in [buttonedit, buttondelete, buttonclearselection]:
    i.config(bg=buttonbg, fg=buttonfg, relief='groove')

accroot.mainloop()
