import tkinter
from tkinter import ttk
import tkcalendar
from ttkwidgets import autocomplete
from tkinter import messagebox
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
import datetime

reproot = tkinter.Tk()
reproot.minsize(width=1150, height=650)
reproot.title('Report Generator')
reproot.iconbitmap(os.path.join('files', 'rep.ico'))
# reproot.focus_set()
reproot.focus_force()

# coloring and styling
winbg = "white smoke"
textfg = "black"

buttonbg = "gainsboro"
buttonfg = "black"

treebg = "white"
treefield = "white"
oddrowbg = "light sky blue"

messageboxbg = "white"
messageboxfg = "black"

style = ttk.Style()
style.theme_use("clam")  # default is vista
style.configure("Treeview", background=treebg, fieldbackground=treefield, rowheight=25, font='TkDefaultFont 9 italic')
style.configure("Treeview.Heading", font='TkDefaultFont 10 bold')
style.map("Treeview", background=[('selected', 'RoyalBlue3')])

reproot.config(bg=winbg)

# variable for storing the data
data = []
gross = 0.00
net = 0.00

item_completion = []
category_completion = []
mode_completion = []

# variables for sorting purpose     None   'a'   'd'
s_date = 'a'
s_item = None
s_category = None
s_mode = None
s_unit = None
s_quantity = None
s_rate = None
s_da = None
s_dr = None
s_net = None

# variable for report directory and file and also data/defaults and files
path_for_report = "Report"
report_file = "ReportGenerated.xlsx"

path_for_defaults = 'Defaults'
category_file = 'category.xlsx'
items_file = 'items.xlsx'
mode_file = 'mode.xlsx'
pair_file = 'pair.xlsx'

path_for_data = 'DataEntry'

# creating folders for data and defaults
for i in [path_for_data, path_for_defaults]:
    if not os.path.exists(i):
        os.mkdir(i)

# creating spreadsheet for the defaults
for i in ['items.xlsx', 'category.xlsx', 'pair.xlsx', 'mode.xlsx', 'unit.xlsx']:
    if not os.path.exists(os.path.join(path_for_defaults, i)):
        workbook = openpyxl.Workbook()
        workbook.worksheets[0].title = 'data'
        workbook.save(os.path.join(path_for_defaults, i))

# Creating folder for saving the report
if not os.path.exists(path_for_report):
    os.mkdir(path_for_report)

# checking any pre existing report file and deleting if any
if os.path.exists(os.path.join(path_for_report, report_file)):
    os.remove(os.path.join(path_for_report, report_file))


# assigning the completion lists
def set_completion():
    global item_completion
    global category_completion
    global mode_completion

    ws = openpyxl.load_workbook(os.path.join(path_for_defaults, items_file))['data']
    for i in ws['A']:
        item_completion.append(i.value)
    if item_completion == [None]:
        item_completion = []
    item_completion.insert(0, 'All')
    del ws

    ws = openpyxl.load_workbook(os.path.join(path_for_defaults, category_file))['data']
    for i in ws['A']:
        category_completion.append(i.value)
    if category_completion == [None]:
        category_completion = []
    category_completion.insert(0, 'All')
    del ws

    ws = openpyxl.load_workbook(os.path.join(path_for_defaults, mode_file))['data']
    for i in ws['A']:
        mode_completion.append(i.value)
    if mode_completion == [None]:
        mode_completion = []
    mode_completion.insert(0, 'All')
    del ws


def permissible_entry(event, wid):
    if wid == itementry:
        if itementry.get() == '':
            itementry.insert(0, item_completion[0])
        else:
            length = len(itementry.get().strip().title())
            itementry.insert(0, itementry.get().strip().title())
            itementry.delete(length, tkinter.END)
        if itementry.get() not in item_completion:
            messagebox.showerror("No Item Found", "The item you have entered is not present")
            itementry.focus_set()

    elif wid == categoryentry:
        if categoryentry.get() == '':
            categoryentry.insert(0, category_completion[0])
        else:
            length = len(categoryentry.get().strip().title())
            categoryentry.insert(0, categoryentry.get().strip().title())
            categoryentry.delete(length, tkinter.END)
        if categoryentry.get() not in category_completion:
            messagebox.showerror("No Category Found", "The category you have entered is not present")
            categoryentry.focus_set()

    elif wid == modeentry:
        if modeentry.get() == '':
            modeentry.insert(0, mode_completion[0])
        else:
            length = len(modeentry.get().strip().title())
            modeentry.insert(0, modeentry.get().strip().title())
            modeentry.delete(length, tkinter.END)
        if modeentry.get() not in mode_completion:
            messagebox.showerror("No Mode Of Payment Found", "The mode of payment you have entered is not present")
            modeentry.focus_set()


def category_auto_fill(event):
    # permissible_entry(event=None, wid=itementry)

    if itementry.get() != item_completion[0]:
        ws = openpyxl.load_workbook(os.path.join(path_for_defaults, pair_file))['data']
        for i in range(1, ws.max_row + 1):
            if ws.cell(column=1, row=i).value == itementry.get():
                categoryentry.delete(0, tkinter.END)
                categoryentry.insert(0, ws.cell(column=2, row=i).value)
                modeentry.focus_set()
                del ws
                break


def on_exit():
    global data
    del data
    if os.path.exists(os.path.join(path_for_report, report_file)):
        os.remove(os.path.join(path_for_report, report_file))
    reproot.destroy()


def validate_to_date(event):
    if not todatepicker.get_date() >= fromdatepicker.get_date():
        messagebox.showerror("Please check the entered Date", "'To Date' must not be a date before the 'From Date'")
        todatepicker.set_date(fromdatepicker.get_date())
        todatepicker.focus_set()


def list_window(a):
    list_data = []

    if a == itementry:
        itementry.focus_set()
    elif a == categoryentry:
        categoryentry.focus_set()

    window = tkinter.Toplevel()
    window.resizable(width=False, height=False)
    window.minsize(width=250, height=300)
    window.config(bg=winbg)
    window.focus_set()

    if a == itementry:
        window.title("Item List")
        list_data = item_completion

    elif a == categoryentry:
        window.title("Category List")
        list_data = category_completion

    def update_list(event):
        b.config(state=tkinter.DISABLED)
        l.delete(0, tkinter.END)
        if e.get() == '':
            for i in list_data:
                l.insert(tkinter.END, i)
        else:
            for i in list_data:
                if e.get().lower() in i.lower():
                    l.insert(tkinter.END, i)

    def on_select(event):
        if l.curselection():
            a.delete(0, tkinter.END)
            a.insert(0, l.get(l.curselection()))
        a.focus_set()
        window.destroy()

    def on_click(event):
        if l.curselection():
            b.config(state=tkinter.NORMAL)
            e.delete(0, tkinter.END)
            e.insert(0, l.get(l.curselection()))

    e = tkinter.Entry(window)
    e.place(relwidth=7 / 10, relx=0.05, rely=0.03, relheight=0.08)
    e.focus_set()
    e.bind('<FocusIn>', update_list)
    e.bind('<KeyRelease>', update_list)

    s = tkinter.Scrollbar(window, orient=tkinter.VERTICAL)
    s.place(relx=0.83, relheight=0.83, rely=0.14)

    l = tkinter.Listbox(window, yscrollcommand=s.set)
    l.place(relx=0.03, rely=0.14, relheight=0.83, relwidth=0.8)
    l.bind('<Double-Button-1>', on_select)
    l.bind('<Return>', on_select)
    l.bind('<ButtonRelease-1>', on_click)
    l.bind('<KeyRelease>', on_click)

    s.config(command=l.yview)

    b = tkinter.Button(window, text='select', command=lambda event=None: on_select(event), state=tkinter.DISABLED)
    b.place(rely=0.03, relwidth=2 / 10, relx=0.77, relheight=0.08)


def tree_update():
    global gross
    global net
    global data

    gross = 0.00
    net = 0.00

    progress = ttk.Progressbar(reproot, mode='indeterminate')
    progress.place(relwidth=0.4025, relx=0.221, rely=0.525)
    progress.start(15)

    tree.delete(*tree.get_children())

    for i in range(len(data)):
        if i % 2 != 0:
            tree.insert(parent='', index=tkinter.END, values=data[i], tags='oddrow')
        else:
            tree.insert(parent='', index=tkinter.END, values=data[i])
        net += float(data[i][9])
        gross += float(data[i][5]) * float(data[i][6])

    progress.place_forget()

    if len(data) == 0:
        tree.insert(parent='', index=tkinter.END, values=('', '', '', '            NO', '       DATA', 'IS          ',
                                                          'AVAILABLE  ', '', '', ''))
    else:
        buttonopen.place(relx=0.87, rely=0.25, relwidth=0.07, relheight=0.055)

    buttonreset.place(relx=0.87, rely=0.40, relwidth=0.07, relheight=0.055)

    labelnumberofentries.config(text='No. of entries = ' + str(len(data)))
    labelnumberofentries.place(relx=0.02, rely=0.92)
    labelgross.config(text='Total Gross = ' + "{:.2f}".format(gross))
    labelgross.place(relx=0.45, rely=0.92)
    labelnet.config(text='Total Net = ' + "{:.2f}".format(net))
    labelnet.place(relx=0.71, rely=0.92)


def generate_report():
    global data

    validate_to_date(None)
    category_auto_fill(None)
    permissible_entry(None, itementry)
    permissible_entry(None, categoryentry)
    permissible_entry(None, modeentry)

    if itementry.get() in item_completion and categoryentry.get() in category_completion\
            and modeentry.get() in mode_completion and  todatepicker.get_date() >= fromdatepicker.get_date():
        for i in [fromdatepicker, todatepicker, itementry, categoryentry, modeentry, buttongenerate, buttonitemlist,
                  buttoncategorylist]:
            i.config(state=tkinter.DISABLED)

        fromdatepicker.focus_set()

        progress = ttk.Progressbar(reproot, mode='indeterminate')
        progress.place(relwidth=0.4025, relx=0.221, rely=0.525)
        progress.start(15)

        from_year = int(fromdatepicker.get()[-4:])
        to_year = int(todatepicker.get()[-4:])

        for year in range(from_year, to_year + 1):
            if os.path.exists(os.path.join(path_for_data, str(year) + '.xlsx')):
                ws = openpyxl.load_workbook(os.path.join(path_for_data, str(year) + '.xlsx'))['data']

                for i in range(2, ws.max_row + 1):
                    if datetime.datetime.strptime(ws.cell(row=i, column=1).value,
                                                  '%d/%m/%Y') >= datetime.datetime.strptime(fromdatepicker.get(),
                                                                                            '%d/%m/%Y'):
                        if datetime.datetime.strptime(ws.cell(row=i, column=1).value,
                                                      '%d/%m/%Y') <= datetime.datetime.strptime(todatepicker.get(),
                                                                                                '%d/%m/%Y'):

                            if itementry.get() == item_completion[0] and categoryentry.get() == category_completion[0] and \
                                    modeentry.get() == mode_completion[0]:
                                d = []
                                for col in range(1, 11):
                                    d.append(ws.cell(row=i, column=col).value)
                                data.append(d)
                                del d

                            else:
                                if itementry.get() != item_completion[0] and categoryentry.get() != category_completion[0] \
                                        and modeentry.get() == mode_completion[0]:
                                    if ws.cell(row=i, column=2).value == itementry.get() and \
                                            ws.cell(row=i, column=3).value == categoryentry.get():
                                        d = []
                                        for col in range(1, 11):
                                            d.append(ws.cell(row=i, column=col).value)
                                        data.append(d)
                                        del d

                                elif itementry.get() == item_completion[0] and categoryentry.get() != category_completion[0] \
                                        and modeentry.get() == mode_completion[0]:
                                    if ws.cell(row=i, column=3).value == categoryentry.get():
                                        d = []
                                        for col in range(1, 11):
                                            d.append(ws.cell(row=i, column=col).value)
                                        data.append(d)
                                        del d

                                elif itementry.get() != item_completion[0] and categoryentry.get() != category_completion[0] \
                                        and modeentry.get() != mode_completion[0]:
                                    if ws.cell(row=i, column=2).value == itementry.get() and \
                                            ws.cell(row=i, column=3).value == categoryentry.get() and \
                                            ws.cell(row=i, column=4).value == modeentry.get():
                                        d = []
                                        for col in range(1, 11):
                                            d.append(ws.cell(row=i, column=col).value)
                                        data.append(d)
                                        del d

                                elif itementry.get() == item_completion[0] and categoryentry.get() != category_completion[0] \
                                        and modeentry.get() != mode_completion[0]:
                                    if ws.cell(row=i, column=3).value == categoryentry.get() and \
                                            ws.cell(row=i, column=4).value == modeentry.get():
                                        d = []
                                        for col in range(1, 11):
                                            d.append(ws.cell(row=i, column=col).value)
                                        data.append(d)
                                        del d

                                elif itementry.get() == item_completion[0] and categoryentry.get() == category_completion[0] \
                                        and modeentry.get() != mode_completion[0]:
                                    if ws.cell(row=i, column=4).value == modeentry.get():
                                        d = []
                                        for col in range(1, 11):
                                            d.append(ws.cell(row=i, column=col).value)
                                        data.append(d)
                                        del d

        progress.place_forget()
        tree_update()


def reset():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    s_date = 'a'
    s_item = None
    s_category = None
    s_mode = None
    s_unit = None
    s_quantity = None
    s_rate = None
    s_da = None
    s_dr = None
    s_net = None

    data.clear()
    tree_update()
    tree.delete(*tree.get_children())

    if os.path.exists(os.path.join(path_for_report, report_file)):
        os.remove(os.path.join(path_for_report, report_file))

    labelnet.place_forget()
    labelnumberofentries.place_forget()
    labelgross.place_forget()

    buttonopen.place_forget()
    buttonreset.place_forget()

    for i in [fromdatepicker, todatepicker, itementry, categoryentry, modeentry, buttongenerate, buttonitemlist,
              buttoncategorylist]:
        i.config(state=tkinter.NORMAL)

    itementry.delete(0, tkinter.END)
    itementry.insert(0, item_completion[0])

    categoryentry.delete(0, tkinter.END)
    categoryentry.insert(0, category_completion[0])

    modeentry.delete(0, tkinter.END)
    modeentry.insert(0, mode_completion[0])

    fromdatepicker.focus_set()


def sort_by_date():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_date is None:
        data.sort(key=lambda x: datetime.datetime.strptime(x[0], '%d/%m/%Y'))
        s_date = 'a'
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_date == 'a':
        data.sort(reverse=True, key=lambda x: datetime.datetime.strptime(x[0], '%d/%m/%Y'))
        s_date = 'd'
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_date == 'd':
        data.sort(key=lambda x: datetime.datetime.strptime(x[0], '%d/%m/%Y'))
        s_date = 'a'
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_item():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_item is None:
        data.sort(key=lambda x: x[1])
        s_date = None
        s_item = 'a'
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_item == 'a':
        data.sort(reverse=True, key=lambda x: x[1])
        s_date = None
        s_item = 'd'
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_item == 'd':
        data.sort(key=lambda x: x[1])
        s_date = None
        s_item = 'a'
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_category():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_category is None:
        data.sort(key=lambda x: x[2])
        s_date = None
        s_item = None
        s_category = 'a'
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_category == 'a':
        data.sort(reverse=True, key=lambda x: x[2])
        s_date = None
        s_item = None
        s_category = 'd'
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_category == 'd':
        data.sort(key=lambda x: x[2])
        s_date = None
        s_item = None
        s_category = 'a'
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_mode():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_mode is None:
        data.sort(key=lambda x: x[3])
        s_date = None
        s_item = None
        s_category = None
        s_mode = 'a'
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_mode == 'a':
        data.sort(reverse=True, key=lambda x: x[3])
        s_date = None
        s_item = None
        s_category = None
        s_mode = 'd'
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_mode == 'd':
        data.sort(key=lambda x: x[3])
        s_date = None
        s_item = None
        s_category = None
        s_mode = 'a'
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_unit():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_unit is None:
        data.sort(key=lambda x: x[4])
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = 'a'
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_unit == 'a':
        data.sort(reverse=True, key=lambda x: x[4])
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = 'd'
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_unit == 'd':
        data.sort(key=lambda x: x[4])
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = 'a'
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_quantity():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_quantity is None:
        data.sort(key=lambda x: float(x[5]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = 'a'
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_quantity == 'a':
        data.sort(reverse=True, key=lambda x: float(x[5]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = 'd'
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    elif s_quantity == 'd':
        data.sort(key=lambda x: float(x[5]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = 'a'
        s_rate = None
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_rate():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_rate is None:
        data.sort(key=lambda x: float(x[6]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = 'a'
        s_da = None
        s_dr = None
        s_net = None

    elif s_rate == 'a':
        data.sort(reverse=True, key=lambda x: float(x[6]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = 'd'
        s_da = None
        s_dr = None
        s_net = None

    elif s_rate == 'd':
        data.sort(key=lambda x: float(x[6]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = 'a'
        s_da = None
        s_dr = None
        s_net = None

    tree_update()


def sort_by_da():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_da is None:
        data.sort(key=lambda x: float(x[7]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = 'a'
        s_dr = None
        s_net = None

    elif s_da == 'a':
        data.sort(reverse=True, key=lambda x: float(x[7]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = 'd'
        s_dr = None
        s_net = None

    elif s_da == 'd':
        data.sort(key=lambda x: float(x[7]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = 'a'
        s_dr = None
        s_net = None

    tree_update()


def sort_by_dr():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_dr is None:
        data.sort(key=lambda x: float(x[8]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = 'a'
        s_net = None

    elif s_dr == 'a':
        data.sort(reverse=True, key=lambda x: float(x[8]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = 'd'
        s_net = None

    elif s_dr == 'd':
        data.sort(key=lambda x: float(x[8]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = 'a'
        s_net = None

    tree_update()


def sort_by_net():
    global s_date
    global s_item
    global s_category
    global s_mode
    global s_unit
    global s_quantity
    global s_rate
    global s_da
    global s_dr
    global s_net

    if s_net is None:
        data.sort(key=lambda x: float(x[9]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = 'a'

    elif s_net == 'a':
        data.sort(reverse=True, key=lambda x: float(x[9]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = 'd'

    elif s_net == 'd':
        data.sort(key=lambda x: float(x[9]))
        s_date = None
        s_item = None
        s_category = None
        s_mode = None
        s_unit = None
        s_quantity = None
        s_rate = None
        s_da = None
        s_dr = None
        s_net = 'a'

    tree_update()


def open_in_app():
    global gross
    global net

    progress = ttk.Progressbar(reproot, mode='indeterminate')
    progress.place(relwidth=0.4025, relx=0.221, rely=0.525)
    progress.start(15)

    if os.path.exists(os.path.join(path_for_report, report_file)):
        os.remove(os.path.join(path_for_report, report_file))

    workbook = openpyxl.Workbook()
    workbook.worksheets[0].title = 'data'
    workbook.save(os.path.join(path_for_report, report_file))

    wb = openpyxl.load_workbook(os.path.join(path_for_report, report_file))
    ws = wb['data']

    ws.cell(row=1, column=1).value = 'DATE'
    ws.cell(row=1, column=2).value = 'ITEM'
    ws.cell(row=1, column=3).value = 'CATEGORY'
    ws.cell(row=1, column=4).value = 'MODE'
    ws.cell(row=1, column=5).value = 'UNIT'
    ws.cell(row=1, column=6).value = 'QUANTITY'
    ws.cell(row=1, column=7).value = 'RATE'
    ws.cell(row=1, column=8).value = 'DIS. ALL'
    ws.cell(row=1, column=9).value = 'DIS. REC'
    ws.cell(row=1, column=10).value = 'NET'

    for i in range(1, 11):  # heading font styling and alignment and border
        ws.cell(row=1, column=i).font = Font(size=11, bold=True)
        ws.cell(row=1, column=i).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=i).border = Border(left=Side(border_style='thick'), right=Side(border_style='thick'),
                                                 top=Side(border_style='thick'), bottom=Side(border_style='thick'))

    for i in range(len(data)):  # inserting the data   and aligning as well and border
        for col in range(1, 11):
            ws.cell(row=2 + i, column=col).value = data[i][col - 1]

            if col == 1:
                ws.cell(row=2 + i, column=col).alignment = Alignment(horizontal='center')
                ws.cell(row=2 + i, column=col).border = Border(left=Side(border_style='thick'))

            if col in [6, 7, 8, 9, 10]:
                ws.cell(row=2 + i, column=col).alignment = Alignment(horizontal='right')

            if col == 10:
                ws.cell(row=2 + i, column=col).border = Border(right=Side(border_style='thick'))

            if i == len(data) - 1:
                if col == 1:
                    ws.cell(row=2 + i, column=col).border = Border(bottom=Side(border_style='thick'),
                                                                   left=Side(border_style='thick'))
                elif col == 10:
                    ws.cell(row=2 + i, column=col).border = Border(bottom=Side(border_style='thick'),
                                                                   right=Side(border_style='thick'))
                else:
                    ws.cell(row=2 + i, column=col).border = Border(bottom=Side(border_style='thick'))

    ws.cell(row=1 + len(data) + 1, column=6).value = 'Gross:'
    ws.cell(row=1 + len(data) + 1, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=1 + len(data) + 1, column=6).font = Font(bold=True)
    ws.cell(row=1 + len(data) + 1, column=6).fill = PatternFill(start_color="F5F51B", fill_type="solid")
    ws.cell(row=1 + len(data) + 1, column=6).border = Border(bottom=Side(border_style='thick'),
                                                             left=Side(border_style='thick'))

    ws.cell(row=1 + len(data) + 1, column=7).value = "{:.2f}".format(gross)
    ws.cell(row=1 + len(data) + 1, column=7).alignment = Alignment(horizontal='right')
    ws.cell(row=1 + len(data) + 1, column=7).font = Font(bold=True)
    ws.cell(row=1 + len(data) + 1, column=7).fill = PatternFill(start_color="F5F51B", fill_type="solid")
    ws.cell(row=1 + len(data) + 1, column=7).border = Border(bottom=Side(border_style='thick'),
                                                             right=Side(border_style='thick'))

    ws.cell(row=1 + len(data) + 1, column=10).value = "{:.2f}".format(net)
    ws.cell(row=1 + len(data) + 1, column=10).alignment = Alignment(horizontal='right')
    ws.cell(row=1 + len(data) + 1, column=10).font = Font(bold=True)
    ws.cell(row=1 + len(data) + 1, column=10).fill = PatternFill(start_color="33FF00", fill_type="solid")
    ws.cell(row=1 + len(data) + 1, column=10).border = Border(bottom=Side(border_style='thick'),
                                                              left=Side(border_style='thick'),
                                                              right=Side(border_style='thick'))

    ws.column_dimensions['A'].width = 13  # specifying the column width
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15

    wb.save(os.path.join(path_for_report, report_file))

    progress.place_forget()

    os.startfile(os.path.join(path_for_report, report_file))


def focus_next(event):
    x = event.widget.tk_focusNext()
    x.focus()
    if x != buttongenerate and x != todatepicker:
        x.select_range(0, tkinter.END)
    return "return"


def mouse_on_open(event):
    info.place(relx=0.87, rely=0.11)
    buttonopen.config(bg="green2")


def mouse_left_open(event):
    info.place_forget()
    buttonopen.config(bg=buttonbg)


set_completion()

labelfromdate = tkinter.Label(reproot, text='From :')
labelfromdate.place(relx=0.02, rely=0.02)
fromdatepicker = tkcalendar.DateEntry(reproot, date_pattern='dd/mm/yyyy')
fromdatepicker.place(relx=0.02, rely=0.05, relwidth=0.133, relheight=0.034)
fromdatepicker.focus_set()
fromdatepicker.bind("<Return>", focus_next)

labeltodate = tkinter.Label(reproot, text='To :')
labeltodate.place(relx=0.173, rely=0.02)
todatepicker = tkcalendar.DateEntry(reproot, date_pattern='dd/mm/yyyy')
todatepicker.place(relx=0.173, rely=0.05, relwidth=0.133, relheight=0.034)
todatepicker.bind("<Return>", focus_next)
todatepicker.bind("<<DateEntrySelected>>", validate_to_date)

labelitems = tkinter.Label(reproot, text='Item :')
labelitems.place(relx=0.326, rely=0.02)
itementry = autocomplete.AutocompleteEntry(reproot, completevalues=item_completion)
itementry.place(relx=0.326, rely=0.05, relwidth=0.133, relheight=0.034)
itementry.insert(0, item_completion[0])
itementry.bind("<Return>", focus_next)
itementry.bind("<FocusOut>", category_auto_fill)

labelcategory = tkinter.Label(reproot, text='Category :')
labelcategory.place(relx=0.509, rely=0.02)
categoryentry = autocomplete.AutocompleteEntry(reproot, completevalues=category_completion)
categoryentry.place(relx=0.509, rely=0.05, relwidth=0.133, relheight=0.034)
categoryentry.insert(0, category_completion[0])
categoryentry.bind("<Return>", focus_next)
# categoryentry.bind("<FocusOut>", lambda event, wid=categoryentry: permissible_entry(event, wid))

labelmode = tkinter.Label(reproot, text='Mode of Payment :')
labelmode.place(relx=0.692, rely=0.02)
modeentry = autocomplete.AutocompleteCombobox(reproot, completevalues=mode_completion)
modeentry.place(relx=0.692, rely=0.05, relwidth=0.133, relheight=0.034)
modeentry.insert(0, mode_completion[0])
modeentry.bind("<Return>", focus_next)
# modeentry.bind("<FocusOut>", lambda event, wid=modeentry: permissible_entry(event, wid))

buttongenerate = tkinter.Button(reproot, text='Generate', relief='groove', command=generate_report)
buttongenerate.place(relx=0.87, rely=0.03, relwidth=0.07, relheight=0.055)
buttongenerate.bind("<Return>", lambda event=None: buttongenerate.invoke())
buttongenerate.bind("<Enter>", lambda event=None: buttongenerate.config(bg="SteelBlue1"))
buttongenerate.bind("<FocusIn>", lambda event=None: buttongenerate.config(bg="SteelBlue1"))
buttongenerate.bind("<Leave>", lambda event=None: buttongenerate.config(bg=buttonbg))
buttongenerate.bind("<FocusOut>", lambda event=None: buttongenerate.config(bg=buttonbg))

buttonopen = tkinter.Button(reproot, text='open', relief='groove', command=open_in_app)
buttonopen.place(relx=0.87, rely=0.25, relwidth=0.07, relheight=0.055)
buttonopen.place_forget()
buttonopen.bind("<Return>", lambda event=None: buttonopen.invoke())
buttonopen.bind("<FocusIn>", lambda event=None: buttonopen.config(bg="green2"))
buttonopen.bind("<FocusOut>", lambda event=None: buttonopen.config(bg=buttonbg))
buttonopen.bind("<Enter>", mouse_on_open)
buttonopen.bind("<Leave>", mouse_left_open)

buttonreset = tkinter.Button(reproot, text='reset', relief='groove', command=reset)
buttonreset.place(relx=0.87, rely=0.40, relwidth=0.07, relheight=0.055)
buttonreset.place_forget()
buttonreset.bind("<Return>", lambda event=None: buttonreset.invoke())
buttonreset.bind("<Enter>", lambda event=None: buttonreset.config(bg="SteelBlue1"))
buttonreset.bind("<FocusIn>", lambda event=None: buttonreset.config(bg="SteelBlue1"))
buttonreset.bind("<Leave>", lambda event=None: buttonreset.config(bg=buttonbg))
buttonreset.bind("<FocusOut>", lambda event=None: buttonreset.config(bg=buttonbg))

buttonexit = tkinter.Button(reproot, text='exit', relief='groove', command=on_exit)
buttonexit.place(relx=0.87, rely=0.55, relwidth=0.07, relheight=0.055)
buttonexit.bind("<Return>", lambda event=None: buttonexit.invoke())
buttonexit.bind("<Enter>", lambda event=None: buttonexit.config(bg="red"))
buttonexit.bind("<FocusIn>", lambda event=None: buttonexit.config(bg="red"))
buttonexit.bind("<Leave>", lambda event=None: buttonexit.config(bg=buttonbg))
buttonexit.bind("<FocusOut>", lambda event=None: buttonexit.config(bg=buttonbg))

buttonitemlist = tkinter.Button(reproot, text='..', command=lambda a=itementry: list_window(a))
buttonitemlist.place(relx=0.469, rely=0.05, relwidth=0.02, relheight=0.034)
buttonitemlist.bind("<Return>", lambda event=None: buttonitemlist.invoke())

buttoncategorylist = tkinter.Button(reproot, text='..', command=lambda a=categoryentry: list_window(a))
buttoncategorylist.place(relx=0.652, rely=0.05, relwidth=0.02, relheight=0.034)
buttoncategorylist.bind("<Return>", lambda event=None: buttoncategorylist.invoke())

labelnumberofentries = tkinter.Label(reproot, text='No. of entries')
labelnumberofentries.place(relx=0.02, rely=0.92)
labelnumberofentries.place_forget()
labelgross = tkinter.Label(reproot, text='Total Gross : ')
labelgross.place(relx=0.422, rely=0.92)
labelgross.place_forget()
labelnet = tkinter.Label(reproot, text='Total Net : ')
labelnet.place(relx=0.7, rely=0.92)
labelnet.place_forget()

scroll = tkinter.Scrollbar(reproot, orient=tkinter.VERTICAL)
scroll.place(relx=0.825, rely=0.13, relheight=0.79)

tree = ttk.Treeview(reproot, show=['headings'], selectmode='browse', yscrollcommand=scroll.set)
tree.place(relx=0.02, rely=0.13, relwidth=0.805, relheight=0.79)

scroll.configure(command=tree.yview)

tree['columns'] = ('date', 'item', 'cat', 'mode', 'unit', 'qty', 'rate', 'da', 'dr', 'net')
reproot.update()
tree.update()
tree.column('date', anchor=tkinter.CENTER, width=int(tree.winfo_width() / 10))
tree.column('item', anchor=tkinter.W, width=int(tree.winfo_width() / 8))
tree.column('cat', anchor=tkinter.W, width=int(tree.winfo_width() / 7.5))
tree.column('mode', anchor=tkinter.W, width=int(tree.winfo_width() / 9))
tree.column('unit', anchor=tkinter.W, width=int(tree.winfo_width() / 12))
tree.column('qty', anchor=tkinter.E, width=int(tree.winfo_width() / 11))
tree.column('rate', anchor=tkinter.E, width=int(tree.winfo_width() / 11))
tree.column('da', anchor=tkinter.E, width=int(tree.winfo_width() / 13))
tree.column('dr', anchor=tkinter.E, width=int(tree.winfo_width() / 13))
tree.column('net', anchor=tkinter.E, width=int(tree.winfo_width() / 8.5))
tree.heading('date', text='Date', command=sort_by_date)  # , anchor=tkinter.W
tree.heading('item', text='Item', command=sort_by_item)
tree.heading('cat', text='Category', command=sort_by_category)
tree.heading('mode', text='Mode', command=sort_by_mode)
tree.heading('unit', text='Unit', command=sort_by_unit)
tree.heading('qty', text='Quantity', command=sort_by_quantity)
tree.heading('rate', text='Rate', command=sort_by_rate)
tree.heading('da', text='Dis. All', command=sort_by_da)
tree.heading('dr', text='Dis. Rec', command=sort_by_dr)
tree.heading('net', text='Net Price', command=sort_by_net)

tree.tag_configure('oddrow', background=oddrowbg)

info = tkinter.Message(reproot, text="Open the current data in your default SpreadSheet Application", bg=messageboxbg,
                       foreground=messageboxfg, font='TkDefaultFont 10 italic')
info.place(relx=0.87, rely=0.11)
info.place_forget()

# styling labels
for i in [labelitems, labelcategory, labelmode, labelfromdate, labeltodate, labelnet, labelgross, labelnumberofentries,
          ]:
    i.config(bg=winbg, fg=textfg, font='TkDefaultFont 10 italic')

# styling buttons
for i in [buttongenerate, buttonopen, buttonreset, buttonexit, buttonitemlist, buttoncategorylist]:
    i.config(bg=buttonbg, fg=buttonfg, font='TkDefaultFont 11 bold italic')

reproot.mainloop()
